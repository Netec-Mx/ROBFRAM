"""Práctica 15 — Sesión 8.

Proceso RPA: lee clientes desde un Excel, transforma los datos (calcula
costo con IVA y clasifica el consumo), y genera un reporte en PDF.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from fpdf import FPDF

IVA_GUATEMALA = 12.0


@dataclass
class Cliente:
    nombre: str
    plan: str
    consumo_gb: float
    costo_base: float
    costo_total: float = 0.0
    clasificacion_consumo: str = ""


def leer_clientes_excel(ruta_excel: str | Path) -> list[Cliente]:
    """Lee la hoja 'Clientes' de un Excel y devuelve una lista de Cliente.

    Raises:
        FileNotFoundError: si la ruta no existe.
    """
    ruta = Path(ruta_excel)
    if not ruta.is_file():
        raise FileNotFoundError(f"No existe el archivo: {ruta}")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    hoja = wb["Clientes"]
    clientes = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        nombre, plan, consumo_gb, costo_base = fila
        clientes.append(Cliente(nombre=nombre, plan=plan, consumo_gb=consumo_gb, costo_base=costo_base))
    return clientes


def clasificar_consumo(consumo_gb: float) -> str:
    """Clasifica el consumo en GB como bajo, medio o alto."""
    if consumo_gb < 50:
        return "bajo"
    if consumo_gb < 100:
        return "medio"
    return "alto"


def transformar_clientes(clientes: list[Cliente]) -> list[Cliente]:
    """Calcula el costo total (con IVA) y clasifica el consumo de cada
    cliente. Devuelve una nueva lista; no muta la lista recibida."""
    transformados = []
    for cliente in clientes:
        costo_total = round(cliente.costo_base * (1 + IVA_GUATEMALA / 100), 2)
        clasificacion = clasificar_consumo(cliente.consumo_gb)
        transformados.append(
            Cliente(
                nombre=cliente.nombre,
                plan=cliente.plan,
                consumo_gb=cliente.consumo_gb,
                costo_base=cliente.costo_base,
                costo_total=costo_total,
                clasificacion_consumo=clasificacion,
            )
        )
    return transformados


def generar_reporte_pdf(clientes: list[Cliente], ruta_salida: str | Path) -> Path:
    """Genera un reporte PDF tabular con los clientes transformados."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Reporte de Clientes - Consumo y Facturación", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)

    columnas = ["Nombre", "Plan", "Consumo GB", "Clasificación", "Costo Total"]
    anchos = [45, 30, 30, 35, 30]
    for col, ancho in zip(columnas, anchos):
        pdf.cell(ancho, 8, col, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 10)
    for cliente in clientes:
        valores = [
            cliente.nombre,
            cliente.plan,
            str(cliente.consumo_gb),
            cliente.clasificacion_consumo,
            f"Q{cliente.costo_total:.2f}",
        ]
        for valor, ancho in zip(valores, anchos):
            pdf.cell(ancho, 8, valor, border=1)
        pdf.ln()

    ruta = Path(ruta_salida)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(ruta))
    return ruta
